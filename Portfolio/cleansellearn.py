from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup as bs
import openpyxl

def init_cdriver():
    options = Options()
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.set_window_size(1920, 1080)
    return driver

def nav_ebs(driver):
    driver.get('https://www.amazon.com/')
    menu = driver.find_element(By.ID, 'nav-hamburger-menu')
    menu.click()
    bests = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//html/body/div[3]/div[2]/div/ul[1]/li[2]/a')))
    bests.click()
    selectlist = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[1]/div[2]/div/div/div/div[2]/div/div[2]/div/div/div[2]/div[20]/a')))
    selectlist.click()
    return driver.current_url

def init_excel():
    xl = openpyxl.Workbook()
    sheet = xl.active
    sheet.title = 'Top 50'
    sheet.append(['Item', 'Link', 'Price'])
    return xl, sheet

def scrapensave(driver, mainlistpage, sheet, i):
    xpath = f"/html/body/div[1]/div[2]/div/div/div[1]/div/div/div[2]/div[1]/div[1]/div[{i}]/div"
    item = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
    item.click()
    page_source = driver.page_source
    soup = bs(page_source, 'html.parser')
    name = soup.find('span', id="productTitle").text.strip()
    price = soup.find('span', {'class': 'a-offscreen'}).text.strip()
    row_num = i + 1
    sheet.cell(row=row_num, column=1).value = name
    sheet.cell(row=row_num, column=3).value = price
    url_cell = sheet.cell(row=row_num, column=2)
    url_cell.hyperlink = driver.current_url
    driver.get(mainlistpage)

def main():
    driver = init_cdriver()
    mainlistpage = nav_ebs(driver)
    xl, sheet = init_excel()

    for i in range(1, 9):
        scrapensave(driver, mainlistpage, sheet, i)

    xl.save('Top 50.xlsx')
    driver.quit()
    
if __name__ == "__main__":
    main()
