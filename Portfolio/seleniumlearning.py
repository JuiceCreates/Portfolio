# Using selenium and beautifulsoup to scrape data from Amazon Best Selling Electronics list and export it to an Excel spreadsheet
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup as bs
import time
import openpyxl

options = Options()
options.add_experimental_option("detach", True) #Added this so browser doesnt close at end of the script. Can be removed.

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

#change window size
driver.set_window_size(1920, 1080) 

#setting the first webpage to be opened
driver.get('https://www.amazon.com/ref=nav_logo') 

#finding expandable menu
menu = driver.find_element(By.ID, 'nav-hamburger-menu') 
menu.click()

#uses xpath to direct the click to "Best Sellers". Amazons dynamic web structure makes it difficult to pinpoint by id or class.
bests = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//html/body/div[3]/div[2]/div/ul[1]/li[2]/a'))) 
bests.click()

#directs to electronics using xpath as well.
selectlist = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[1]/div[2]/div/div/div/div[2]/div/div[2]/div/div/div[2]/div[20]/a')))
selectlist.click()

#saves the electronics best sellers list to mainlistpage
MAINLISTPAGE = driver.current_url 

#this opens the new Excel workbook, names the sheet and the column titles
xl = openpyxl.Workbook() 
sheet = xl.active
sheet.title = 'Top 50'
sheet.append(['Item', 'Link', 'Price'])

#for loop iterating through levels of divs on the webpage. It grabs the name, link, and price and exports it to Excel 
for i in range(1,9):
    
    xpath = f"/html/body/div[1]/div[2]/div/div/div[1]/div/div/div[2]/div[1]/div[1]/div[{i}]/div"
    item = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
    item.click()
    url = driver.current_url
    page_source = driver.page_source
    soup = bs(page_source, 'html.parser')
    name = soup.find('span', id="productTitle").text.strip()
    price = soup.find('span', {'class': 'a-offscreen'}).text.strip()
    row_num = i + 1 
    sheet.cell(row=row_num, column=1).value = name
    sheet.cell(row=row_num, column=3).value = price
    urlcell = sheet.cell(row=row_num, column=2)  
    urlcell.hyperlink = url 

    driver.get(MAINLISTPAGE)

xl.save('Top 50.xlsx')